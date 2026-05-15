"""
Modern Grammar Quiz App
BS AI — COMSATS University Lahore
By Muhammad Shoaib Tahir
"""

import streamlit as st
import hashlib
import json
import time
import random
import os
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Persistent storage helpers ───────────────────────────────────────────────
# Results are saved to a local JSON file so they survive server restarts.
RESULTS_FILE = "quiz_results.json"

def load_all_results() -> dict:
    """Load all student results from disk."""
    if os.path.exists(RESULTS_FILE):
        try:
            with open(RESULTS_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_result(roll: str, name: str, quiz_name: str, score_data: dict):
    """Persist a student's quiz result to disk."""
    all_results = load_all_results()
    if roll not in all_results:
        all_results[roll] = {"name": name, "roll": roll, "quizzes": {}}
    # Strip non-serialisable objects (question dicts with lists etc.)
    safe_detail = []
    for item in score_data.get("detail", []):
        safe_detail.append({
            "q_text": item["q"]["text"],
            "q_type": item["q"]["type"],
            "ans":    item["ans"],
            "score":  item["score"],
            "max":    item["max"],
        })
    all_results[roll]["quizzes"][quiz_name] = {
        "auto_score":    score_data["auto_score"],
        "auto_total":    score_data["auto_total"],
        "elapsed":       score_data.get("elapsed", 0),
        "submitted_at":  score_data.get("submitted_at", ""),
        "detail":        safe_detail,
    }
    with open(RESULTS_FILE, "w") as f:
        json.dump(all_results, f, indent=2)

def is_quiz_locked(roll: str, quiz_name: str) -> bool:
    """Return True if this student already submitted this quiz."""
    all_results = load_all_results()
    return roll in all_results and quiz_name in all_results[roll].get("quizzes", {})

def generate_excel() -> bytes:
    """Build a styled Excel report of all student results and return as bytes."""
    all_results = load_all_results()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quiz Results"

    # ── Colour palette ───────────────────────────────────────────────────
    HDR_BG   = PatternFill("solid", fgColor="0D1B2A")
    HDR_FG   = Font(bold=True, color="E8A020", size=11)
    SUB_BG   = PatternFill("solid", fgColor="1B7F79")
    SUB_FG   = Font(bold=True, color="FFFFFF", size=10)
    ALT_BG   = PatternFill("solid", fgColor="EAF6F6")
    GOOD_BG  = PatternFill("solid", fgColor="D5F5E3")
    BAD_BG   = PatternFill("solid", fgColor="FADBD8")
    MID_BG   = PatternFill("solid", fgColor="FEF9E7")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    QUIZ_NAMES = list(QUIZZES.keys())
    Q1 = QUIZ_NAMES[0] if len(QUIZ_NAMES) > 0 else "Quiz 1"
    Q2 = QUIZ_NAMES[1] if len(QUIZ_NAMES) > 1 else "Quiz 2"

    # ── Title row ────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Modern Grammar Quiz — BS AI SP26 | COMSATS University Lahore"
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = HDR_BG
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ── Sub-header ───────────────────────────────────────────────────────
    ws.merge_cells("A2:L2")
    sub = ws["A2"]
    sub.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  By Muhammad Shoaib Tahir"
    sub.font  = Font(italic=True, color="555555", size=10)
    sub.alignment = center
    ws.row_dimensions[2].height = 18

    # ── Column headers ───────────────────────────────────────────────────
    headers = [
        "S.No", "Roll Number", "Student Name",
        f"Quiz 1\nMCQ+Fill\n(/13)", f"Quiz 1\nShort Q\n(/4 manual)", f"Quiz 1\nTime",
        f"Quiz 2\nMCQ+Fill\n(/13)", f"Quiz 2\nShort Q\n(/4 manual)", f"Quiz 2\nTime",
        "Total Auto\n(/26)", "Grade", "Status"
    ]
    ws.row_dimensions[3].height = 40
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_i, value=h)
        cell.font      = SUB_FG
        cell.fill      = SUB_BG
        cell.alignment = center
        cell.border    = border

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = [6, 20, 28, 14, 16, 10, 14, 16, 10, 12, 8, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ────────────────────────────────────────────────────────
    sno = 1
    for roll, sdata in sorted(all_results.items()):
        name     = sdata.get("name", "")
        quizzes  = sdata.get("quizzes", {})

        q1_data  = quizzes.get(Q1, {})
        q2_data  = quizzes.get(Q2, {})

        q1_auto  = q1_data.get("auto_score", "—")
        q2_auto  = q2_data.get("auto_score", "—")

        # count short answers
        def count_short(qdata):
            return sum(1 for d in qdata.get("detail", []) if d["q_type"] == "short")

        q1_short_count = count_short(q1_data)
        q2_short_count = count_short(q2_data)

        def fmt_elapsed(sec):
            if not sec: return "—"
            m, s = divmod(int(sec), 60)
            return f"{m:02d}:{s:02d}"

        q1_time  = fmt_elapsed(q1_data.get("elapsed"))
        q2_time  = fmt_elapsed(q2_data.get("elapsed"))

        total_auto = 0
        if isinstance(q1_auto, (int, float)): total_auto += q1_auto
        if isinstance(q2_auto, (int, float)): total_auto += q2_auto
        total_str = f"{total_auto:.1f}" if (isinstance(q1_auto,(int,float)) or isinstance(q2_auto,(int,float))) else "—"

        pct = (total_auto / 26 * 100) if isinstance(total_auto, (int,float)) else 0
        grade = "A+" if pct>=90 else "A" if pct>=80 else "B" if pct>=70 else "C" if pct>=60 else "D" if pct>=50 else "F"
        both_done = bool(q1_data) and bool(q2_data)
        status = "✅ Both Done" if both_done else ("⏳ Quiz 1 Only" if q1_data else ("⏳ Quiz 2 Only" if q2_data else "❌ Not Started"))

        row_vals = [
            sno, roll.upper(), name,
            f"{q1_auto:.1f}" if isinstance(q1_auto, (int,float)) else "—",
            f"— / {q1_short_count*2} pending" if q1_short_count else "—",
            q1_time,
            f"{q2_auto:.1f}" if isinstance(q2_auto, (int,float)) else "—",
            f"— / {q2_short_count*2} pending" if q2_short_count else "—",
            q2_time,
            total_str, grade, status
        ]

        row_num  = sno + 3
        bg_fill  = ALT_BG if sno % 2 == 0 else None
        grade_fill = GOOD_BG if pct >= 70 else (MID_BG if pct >= 50 else BAD_BG)

        for col_i, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_num, column=col_i, value=val)
            cell.alignment = center if col_i != 3 else left
            cell.border    = border
            if col_i == 11:   # Grade column
                cell.fill = grade_fill
                cell.font = Font(bold=True, size=11)
            elif bg_fill:
                cell.fill = bg_fill

        ws.row_dimensions[row_num].height = 20
        sno += 1

    # ── Summary stats at bottom ──────────────────────────────────────────
    blank_row = sno + 4
    ws.cell(row=blank_row, column=1, value="Summary").font = Font(bold=True, size=11, color="0D1B2A")
    ws.cell(row=blank_row+1, column=1, value="Total Students Registered:").font = Font(bold=True)
    ws.cell(row=blank_row+1, column=2, value=len(STUDENT_LIST))
    ws.cell(row=blank_row+2, column=1, value="Students Submitted Both Quizzes:").font = Font(bold=True)
    both = sum(1 for v in all_results.values() if Q1 in v.get("quizzes",{}) and Q2 in v.get("quizzes",{}))
    ws.cell(row=blank_row+2, column=2, value=both)
    ws.cell(row=blank_row+3, column=1, value="Students Not Started:").font = Font(bold=True)
    ws.cell(row=blank_row+3, column=2, value=len(STUDENT_LIST) - len(all_results))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── Page config ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Modern Grammar Quiz | BS AI",
    page_icon="📘",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;900&family=Lato:wght@300;400;700&display=swap');

:root {
    --navy:    #0d1b2a;
    --teal:    #1b7f79;
    --gold:    #e8a020;
    --cream:   #f5f0e8;
    --white:   #ffffff;
    --shadow:  0 4px 24px rgba(13,27,42,0.13);
}

html, body, [class*="css"]  {
    font-family: 'Lato', sans-serif;
    background-color: var(--cream);
    color: var(--navy);
}

/* Header banner */
.hero-banner {
    background: linear-gradient(135deg, #0d1b2a 0%, #1b7f79 100%);
    border-radius: 16px;
    padding: 2.5rem 2rem 2rem;
    margin-bottom: 2rem;
    text-align: center;
    box-shadow: var(--shadow);
}
.hero-banner h1 {
    font-family: 'Playfair Display', serif;
    color: var(--gold);
    font-size: 2.6rem;
    margin: 0 0 0.3rem;
    letter-spacing: 0.5px;
}
.hero-banner p {
    color: rgba(255,255,255,0.82);
    font-size: 1rem;
    margin: 0;
    font-weight: 300;
}
.hero-banner .badge {
    display: inline-block;
    background: var(--gold);
    color: var(--navy);
    font-weight: 700;
    font-size: 0.75rem;
    padding: 3px 12px;
    border-radius: 20px;
    letter-spacing: 1px;
    margin-bottom: 0.8rem;
    text-transform: uppercase;
}

/* Card containers */
.card {
    background: #ffffff !important;
    border-radius: 14px;
    padding: 2rem;
    margin-bottom: 1.5rem;
    box-shadow: var(--shadow);
    border-left: 5px solid var(--teal);
    color: #0d1b2a !important;
}
.card * { color: #0d1b2a !important; }

/* Question number chip */
.q-chip {
    background: var(--teal);
    color: white;
    font-weight: 700;
    font-size: 0.78rem;
    padding: 3px 12px;
    border-radius: 20px;
    display: inline-block;
    margin-bottom: 0.5rem;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}
.q-type-fill  { background: #6a4c93; }
.q-type-short { background: #c0392b; }
.q-type-mcq   { background: var(--teal); }

/* Question text */
.q-text {
    font-family: 'Playfair Display', serif;
    font-size: 1.12rem;
    color: #0d1b2a !important;
    margin: 0.4rem 0 1rem;
    line-height: 1.55;
}

/* Progress bar */
.prog-wrap {
    background: #dde;
    border-radius: 8px;
    height: 10px;
    margin: 0.8rem 0;
    overflow: hidden;
}
.prog-fill {
    height: 100%;
    background: linear-gradient(90deg, var(--teal), var(--gold));
    border-radius: 8px;
    transition: width 0.4s ease;
}

/* Score box */
.score-box {
    background: linear-gradient(135deg, #0d1b2a, #1b7f79);
    border-radius: 16px;
    padding: 2.5rem;
    text-align: center;
    color: white;
    box-shadow: var(--shadow);
}
.score-box .big-score {
    font-family: 'Playfair Display', serif;
    font-size: 5rem;
    color: var(--gold);
    line-height: 1;
}
.score-box .score-label {
    font-size: 1.1rem;
    opacity: 0.85;
    margin-top: 0.4rem;
}

/* Answer review */
.ans-correct { border-left: 4px solid #27ae60; background:#f0fff4 !important; padding:0.8rem 1rem; border-radius:8px; margin:0.5rem 0; color:#0d1b2a !important; }
.ans-wrong   { border-left: 4px solid #e74c3c; background:#fff5f5 !important; padding:0.8rem 1rem; border-radius:8px; margin:0.5rem 0; color:#0d1b2a !important; }
.ans-partial { border-left: 4px solid #f39c12; background:#fffbf0 !important; padding:0.8rem 1rem; border-radius:8px; margin:0.5rem 0; color:#0d1b2a !important; }
.ans-correct b, .ans-wrong b, .ans-partial b { color: #0d1b2a !important; }

/* Button tweaks */
.stButton > button {
    background: linear-gradient(135deg, var(--teal), #0d6b66);
    color: white;
    border: none;
    border-radius: 8px;
    font-weight: 700;
    font-size: 1rem;
    padding: 0.65rem 2rem;
    transition: all 0.2s;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #1d9990, #0d1b2a);
    transform: translateY(-1px);
    box-shadow: 0 6px 18px rgba(27,127,121,0.35);
}

/* Input fields */
.stTextInput > div > div > input {
    border-radius: 8px;
    border: 2px solid #ccc;
    font-size: 1rem;
    padding: 0.5rem 0.8rem;
}
.stTextInput > div > div > input:focus {
    border-color: var(--teal);
    box-shadow: 0 0 0 3px rgba(27,127,121,0.15);
}

/* Warning modal */
.confirm-box {
    background: #2a1a00 !important;
    border: 2px solid #e8a020;
    border-radius: 12px;
    padding: 1.5rem;
    text-align: center;
    color: #ffe0a0 !important;
}
.confirm-box h3 {
    color: #ffc44d !important;
    font-size: 1.2rem;
    margin-bottom: 0.5rem;
}
.confirm-box p {
    color: #ffe0a0 !important;
    font-size: 0.97rem;
    line-height: 1.6;
}
.confirm-box strong {
    color: #ffd166 !important;
}

/* Quiz selector tabs */
.stTabs [data-baseweb="tab"] {
    font-weight: 700;
    font-size: 1rem;
}

/* Sidebar */
.sidebar-info {
    background: var(--navy);
    color: white;
    border-radius: 10px;
    padding: 1rem;
    font-size: 0.9rem;
    line-height: 1.7;
}

/* Footer */
.footer {
    text-align: center;
    color: #888;
    font-size: 0.82rem;
    margin-top: 3rem;
    padding-top: 1rem;
    border-top: 1px solid #ddd;
}
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
#  QUESTION BANK  (2 quizzes × 15 questions each)
# ═══════════════════════════════════════════════════════════════════════════

QUIZ_1_QUESTIONS = [
    # ── MCQs (Q1–10) ──────────────────────────────────────────────────────
    {
        "id": "q1_1", "type": "mcq",
        "text": "Which of the following is a COMPLETE sentence?",
        "options": ["Running fast.", "She sings beautifully.", "The children.", "In the garden."],
        "answer": "She sings beautifully.",
        "topic": "Unit 1",
    },
    {
        "id": "q1_2", "type": "mcq",
        "text": "An imperative sentence is used to:",
        "options": ["Ask a question", "Share information", "Give a command or request", "Show strong emotion"],
        "answer": "Give a command or request",
        "topic": "Unit 1",
    },
    {
        "id": "q1_3", "type": "mcq",
        "text": "In the sentence 'She gave me a book,' what is the indirect object?",
        "options": ["She", "me", "book", "gave"],
        "answer": "me",
        "topic": "Unit 2",
    },
    {
        "id": "q1_4", "type": "mcq",
        "text": "A phrase is different from a clause because a phrase:",
        "options": [
            "Has a subject and verb",
            "Does NOT have both a subject and a verb",
            "Always starts with a preposition",
            "Expresses a complete thought",
        ],
        "answer": "Does NOT have both a subject and a verb",
        "topic": "Unit 3",
    },
    {
        "id": "q1_5", "type": "mcq",
        "text": "Which sentence correctly uses Subject-Verb Agreement?",
        "options": [
            "The cats runs fast.",
            "Everyone are happy.",
            "Neither the teacher nor the students are ready.",
            "Ali and Sara is friends.",
        ],
        "answer": "Neither the teacher nor the students are ready.",
        "topic": "Unit 6",
    },
    {
        "id": "q1_6", "type": "mcq",
        "text": "'What a beautiful flower!' — What type of sentence is this?",
        "options": ["Declarative", "Interrogative", "Imperative", "Exclamatory"],
        "answer": "Exclamatory",
        "topic": "Unit 1",
    },
    {
        "id": "q1_7", "type": "mcq",
        "text": "In the sentence 'Ali plays cricket every day,' the predicate is:",
        "options": ["Ali", "plays cricket", "plays cricket every day", "cricket every day"],
        "answer": "plays cricket every day",
        "topic": "Unit 2",
    },
    {
        "id": "q1_8", "type": "mcq",
        "text": "Which of the following is a prepositional phrase?",
        "options": ["very quickly", "the big red apple", "under the bed", "is eating"],
        "answer": "under the bed",
        "topic": "Unit 3",
    },
    {
        "id": "q1_9", "type": "mcq",
        "text": "Transformation of 'Ali wrote a letter' into passive voice gives:",
        "options": [
            "Ali is writing a letter.",
            "A letter was written by Ali.",
            "A letter is writing by Ali.",
            "Ali had written a letter.",
        ],
        "answer": "A letter was written by Ali.",
        "topic": "Unit 4",
    },
    {
        "id": "q1_10", "type": "mcq",
        "text": "Which sentence shows correct inversion after a negative word?",
        "options": [
            "Rarely he comes on time.",
            "Rarely does he come on time.",
            "He rarely does comes on time.",
            "Does rarely he come on time.",
        ],
        "answer": "Rarely does he come on time.",
        "topic": "Unit 4",
    },
    # ── Fill in the blanks (Q11–13) ────────────────────────────────────────
    {
        "id": "q1_11", "type": "fill",
        "text": "A sentence must have a ___________ (who/what) and a ___________ (action or state).",
        "answer": ["subject", "verb"],
        "blanks": 2,
        "topic": "Unit 1",
    },
    {
        "id": "q1_12", "type": "fill",
        "text": "An independent clause can stand ___________ on its own.",
        "answer": ["alone"],
        "blanks": 1,
        "topic": "Unit 8",
    },
    {
        "id": "q1_13", "type": "fill",
        "text": "In 'He sent Sara a letter,' Sara is the ___________ object and letter is the ___________ object.",
        "answer": ["indirect", "direct"],
        "blanks": 2,
        "topic": "Unit 2",
    },
    # ── Short questions (Q14–15) ───────────────────────────────────────────
    {
        "id": "q1_14", "type": "short",
        "text": "What is a Subject Complement? Give one example.",
        "answer": "A subject complement comes after a linking verb (is, are, was, seem, look, become) and describes the subject. Example: 'He looks tired.' — 'tired' is the subject complement.",
        "topic": "Unit 2",
    },
    {
        "id": "q1_15", "type": "short",
        "text": "Write the rule for Subject-Verb Agreement when two subjects are joined by 'and'.",
        "answer": "When two subjects are joined by 'and', use a plural verb. Example: Ali and Sara are friends. Exception: fixed pairs like 'bread and butter' take a singular verb.",
        "topic": "Unit 6",
    },
]

QUIZ_2_QUESTIONS = [
    # ── MCQs (Q1–10) ──────────────────────────────────────────────────────
    {
        "id": "q2_1", "type": "mcq",
        "text": "Which is a subordinating conjunction used to show REASON?",
        "options": ["although", "when", "because", "if"],
        "answer": "because",
        "topic": "Unit 5",
    },
    {
        "id": "q2_2", "type": "mcq",
        "text": "'What she said surprised me.' — What type of clause is 'What she said'?",
        "options": ["Adjective clause", "Adverb clause", "Noun clause", "Independent clause"],
        "answer": "Noun clause",
        "topic": "Unit 8",
    },
    {
        "id": "q2_3", "type": "mcq",
        "text": "Which method of synthesis uses a relative pronoun?",
        "options": [
            "Using a coordinating conjunction",
            "Using a participle",
            "Using a relative pronoun (who, which, that)",
            "Using a gerund",
        ],
        "answer": "Using a relative pronoun (who, which, that)",
        "topic": "Unit 7",
    },
    {
        "id": "q2_4", "type": "mcq",
        "text": "In 'She painted the wall blue,' 'blue' is called:",
        "options": ["Direct object", "Indirect object", "Object complement", "Subject complement"],
        "answer": "Object complement",
        "topic": "Unit 2",
    },
    {
        "id": "q2_5", "type": "mcq",
        "text": "Which of the following sentences has a VERB PHRASE?",
        "options": [
            "the big red apple",
            "has been studying",
            "full of energy",
            "in the morning",
        ],
        "answer": "has been studying",
        "topic": "Unit 3",
    },
    {
        "id": "q2_6", "type": "mcq",
        "text": "Choose the sentence with the correct use of 'There is / There are':",
        "options": [
            "There is cats on the roof.",
            "There are a cat on the roof.",
            "There are cats on the roof.",
            "There is cats on the roof.",
        ],
        "answer": "There are cats on the roof.",
        "topic": "Unit 6",
    },
    {
        "id": "q2_7", "type": "mcq",
        "text": "Synthesis by participle: 'She opened the door. She walked in.' becomes:",
        "options": [
            "She opened the door and walked in.",
            "Opening the door, she walked in.",
            "Because she opened the door, she walked in.",
            "She walked in when opening the door.",
        ],
        "answer": "Opening the door, she walked in.",
        "topic": "Unit 7",
    },
    {
        "id": "q2_8", "type": "mcq",
        "text": "Which clause type describes a noun and begins with who/which/that?",
        "options": ["Noun clause", "Adverb clause", "Adjective (Relative) clause", "Independent clause"],
        "answer": "Adjective (Relative) clause",
        "topic": "Unit 8",
    },
    {
        "id": "q2_9", "type": "mcq",
        "text": "'I know that he is honest.' — The bold clause acts as:",
        "options": ["Subject", "Object", "Adverb", "Complement of predicate"],
        "answer": "Object",
        "topic": "Unit 8",
    },
    {
        "id": "q2_10", "type": "mcq",
        "text": "Sentence: 'Although she was tired, she finished her homework.' — The main clause is:",
        "options": [
            "Although she was tired",
            "she finished her homework",
            "although",
            "she was tired",
        ],
        "answer": "she finished her homework",
        "topic": "Unit 5",
    },
    # ── Fill in the blanks (Q11–13) ────────────────────────────────────────
    {
        "id": "q2_11", "type": "fill",
        "text": "In inversion after 'So', the structure is: So ___________ I (to agree with 'She is tired').",
        "answer": ["am"],
        "blanks": 1,
        "topic": "Unit 4",
    },
    {
        "id": "q2_12", "type": "fill",
        "text": "Indefinite pronouns like 'everyone', 'nobody', and 'each' always take a ___________ verb.",
        "answer": ["singular"],
        "blanks": 1,
        "topic": "Unit 6",
    },
    {
        "id": "q2_13", "type": "fill",
        "text": "An adverb clause that tells time uses conjunctions like when, while, after, before, and ___________.",
        "answer": ["since", "until"],
        "blanks": 1,
        "topic": "Unit 8",
    },
    # ── Short questions (Q14–15) ───────────────────────────────────────────
    {
        "id": "q2_14", "type": "short",
        "text": "Explain Transformation with one example of Active to Passive voice.",
        "answer": "Transformation means changing a sentence from one form to another without changing its meaning. Active: 'She baked a cake.' → Passive: 'A cake was baked by her.'",
        "topic": "Unit 4",
    },
    {
        "id": "q2_15", "type": "short",
        "text": "What is synthesis of sentences? Give one example using a coordinating conjunction.",
        "answer": "Synthesis means joining two or more simple sentences into one sentence without changing the meaning. Example: 'He was tired. He kept working.' → 'He was tired, but he kept working.'",
        "topic": "Unit 7",
    },
]

QUIZZES = {
    "Quiz 1 — Sentence Structure, Phrases & Agreement": QUIZ_1_QUESTIONS,
    "Quiz 2 — Complex Sentences, Clauses & Synthesis": QUIZ_2_QUESTIONS,
}

STUDENT_LIST = {
    "sp26-bai-001": "AAMNA BASIT", "sp26-bai-002": "ABDUL HANAN", "sp26-bai-003": "ABDUL REHMAN AZAM",
    "sp26-bai-004": "ABDUL REHMAN BASIT", "sp26-bai-005": "ABEER AMINA", "sp26-bai-006": "AHMAD AFZAL",
    "sp26-bai-007": "AHMAD SHAHZAD", "sp26-bai-008": "AHMAD TALAL", "sp26-bai-009": "ALEEHA KHURRAM",
    "sp26-bai-010": "ALI ISHTIAQ", "sp26-bai-011": "ALISHBA JAVED", "sp26-bai-012": "ALIZA CHOUDHARY",
    "sp26-bai-013": "AMAMA AKBAR", "sp26-bai-014": "AMNA KASHIF", "sp26-bai-015": "ARSLAN MUMTAZ",
    "sp26-bai-016": "ATIKA MEHRBAN", "sp26-bai-017": "EMAN SHAFAQAT", "sp26-bai-018": "FAIZA SHAKEEL",
    "sp26-bai-019": "FARYAL SARFRAZ", "sp26-bai-020": "FATIMA ANSAR", "sp26-bai-021": "FATIMA ASHFAQ",
    "sp26-bai-022": "FATIMA ZAHRA", "sp26-bai-023": "HAMNA KHUBAIB", "sp26-bai-024": "HANIA",
    "sp26-bai-025": "IQRA NAVEED", "sp26-bai-026": "MUHAMMAD ABDULLAH BILAL", "sp26-bai-027": "MAHAM SHAHZAD",
    "sp26-bai-028": "MAHEEN", "sp26-bai-029": "MANAHIL TARIQ", "sp26-bai-030": "MARYAM ASHFAQ",
    "sp26-bai-031": "MARYAM MUMTAZ", "sp26-bai-032": "MINAHIL SEHAR", "sp26-bai-033": "MUHAMMAD ADNAN",
    "sp26-bai-034": "MUHAMMAD ARMAN HAMZA", "sp26-bai-035": "MUHAMMAD IBRAHIM", "sp26-bai-036": "MUHAMMAD SAOUD AZIZ",
    "sp26-bai-037": "MUHAMMAD SHERAZ BHATTI", "sp26-bai-038": "MUHAMMAD TALHA WAQAR",
    "sp26-bai-039": "MUHAMMAD AWAIS ZAHOOR", "sp26-bai-040": "NAIK BAKHAT IRFAN",
    "sp26-bai-041": "QIRAT HAFSA", "sp26-bai-042": "RANA ZUNNURAIN ZAHID", "sp26-bai-043": "ROSHAN AHMED",
    "sp26-bai-044": "SARA ASGHAR", "sp26-bai-045": "SHARMEEN IMTIAZ", "sp26-bai-046": "SWERA KAMRAN",
    "sp26-bai-047": "SYED ABDUL BARI", "sp26-bai-048": "SYED ZARYAB HAIDER", "sp26-bai-049": "TAHA RIZWAN",
    "sp26-bai-050": "TEHREEM SHAHID", "sp26-bai-051": "USAMA NASIR", "sp26-bai-052": "ZAINAB HAROON",
    "sp26-bai-053": "ZAINAB JAVED", "sp26-bai-054": "ZAINAB SHAHZAD", "sp26-bai-055": "AYISHA",
    "sp26-bai-056": "GHANIA ASHRAF",
}


# ═══════════════════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════

def get_seed(roll: str, quiz_name: str) -> int:
    """Deterministic seed per student per quiz — same questions, shuffled order."""
    raw = f"{roll.lower().strip()}|{quiz_name}"
    return int(hashlib.md5(raw.encode()).hexdigest(), 16) % (2**31)


def shuffle_questions(questions: list, seed: int) -> list:
    """Return questions in a consistent shuffled order for this student."""
    rng = random.Random(seed)
    q_copy = [q.copy() for q in questions]
    # shuffle MCQ options deterministically too
    for q in q_copy:
        if q["type"] == "mcq":
            opts = q["options"][:]
            rng.shuffle(opts)
            q["options"] = opts
    rng.shuffle(q_copy)
    return q_copy


def score_mcq(question: dict, answer: str) -> float:
    return 1.0 if answer.strip() == question["answer"].strip() else 0.0


def score_fill(question: dict, answers: list) -> float:
    correct_answers = question["answer"]
    total = len(correct_answers)
    got = 0
    for i, ans in enumerate(answers):
        if i < total and ans.strip().lower() == correct_answers[i].strip().lower():
            got += 1
    return got / total if total > 0 else 0.0


def progress_bar(current: int, total: int):
    pct = int((current / total) * 100)
    st.markdown(
        f'<div class="prog-wrap"><div class="prog-fill" style="width:{pct}%"></div></div>',
        unsafe_allow_html=True,
    )
    st.caption(f"Question {current} of {total} — {pct}% complete")


def type_chip(q_type: str):
    labels = {"mcq": "Multiple Choice", "fill": "Fill in the Blank", "short": "Short Answer"}
    css    = {"mcq": "q-type-mcq", "fill": "q-type-fill", "short": "q-type-short"}
    return f'<span class="q-chip {css[q_type]}">{labels[q_type]}</span>'


# ═══════════════════════════════════════════════════════════════════════════
#  SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════════════════

defaults = {
    "page": "login",           # login | quiz_select | quiz | result
    "student_name": "",
    "student_roll": "",
    "active_quiz": None,
    "shuffled_qs": [],
    "answers": {},
    "q_index": 0,
    "submitted": False,
    "scores": {},
    "confirm_submit": False,
    "start_time": None,
    "completed_quizzes": [],
    "locked_quizzes": [],      # quizzes locked from disk (one-time login)
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ═══════════════════════════════════════════════════════════════════════════
#  HEADER (always visible)
# ═══════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="hero-banner">
  <div class="badge">COMSATS University Lahore — SP26</div>
  <h1>📘 Modern Grammar Quiz</h1>
  <p>BS Artificial Intelligence &nbsp;•&nbsp; English Language Enhancement &nbsp;•&nbsp; By Muhammad Shoaib Tahir</p>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
#  PAGE: LOGIN
# ═══════════════════════════════════════════════════════════════════════════

if st.session_state.page == "login":
    st.markdown("### 👤 Student Login")
    st.markdown("Please enter your details to begin the quiz.")

    col1, col2 = st.columns(2)
    with col1:
        name_input = st.text_input("Full Name", placeholder="e.g. FATIMA ANSAR")
    with col2:
        roll_input = st.text_input("Roll Number (last part)", placeholder="e.g. sp26-bai-020")

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("🚀 Start Quiz Portal", use_container_width=True):
        name_clean = name_input.strip().upper()
        roll_clean = roll_input.strip().lower()
        if not name_clean or not roll_clean:
            st.error("⚠️ Please enter both your name and roll number.")
        elif roll_clean not in STUDENT_LIST:
            st.warning(
                "⚠️ Roll number not found in class list. "
                "Please double-check (e.g. sp26-bai-001). "
                "Contact your instructor if the issue persists."
            )
        else:
            expected_name = STUDENT_LIST[roll_clean]
            if name_clean.split()[0] not in expected_name:
                st.error(
                    "⚠️ Name does not match roll number. "
                    "Expected name starting with the registered first name. "
                    "Please verify your credentials."
                )
            else:
                # ── Load any previously saved results for this student ──
                all_results = load_all_results()
                student_saved = all_results.get(roll_clean, {}).get("quizzes", {})
                locked = [qn for qn in QUIZZES if qn in student_saved]

                st.session_state.student_name      = expected_name
                st.session_state.student_roll      = roll_clean
                st.session_state.completed_quizzes = locked
                st.session_state.locked_quizzes    = locked
                # Restore score summaries so result page works after re-login
                for qn in locked:
                    st.session_state.scores[qn] = student_saved[qn]
                st.session_state.page = "quiz_select"
                st.rerun()

    st.markdown("""
<div style="margin-top:2rem; padding:1.2rem 1.4rem; background:#1b4f6b; border-radius:10px;
            font-size:0.92rem; color:#ffffff; border-left:4px solid #5bc8e8;">
<span style="font-weight:700; font-size:1rem; color:#5bc8e8;">ℹ️ Instructions</span><br><br>
<span style="color:#e0f4ff;">• Each quiz has <b>15 questions</b>: 10 MCQs, 3 Fill in the Blanks, 2 Short Questions.</span><br>
<span style="color:#e0f4ff;">• Questions are shuffled uniquely per student but remain the same each time you return.</span><br>
<span style="color:#e0f4ff;">• You <b>cannot</b> re-attempt a completed quiz.</span><br>
<span style="color:#e0f4ff;">• A confirmation dialog will appear before final submission.</span><br>
<span style="color:#e0f4ff;">• MCQs and Fill-in-the-blanks are auto-graded. Short answers are reviewed manually.</span>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
#  PAGE: QUIZ SELECT
# ═══════════════════════════════════════════════════════════════════════════

elif st.session_state.page == "quiz_select":
    st.markdown(f"### 👋 Welcome, **{st.session_state.student_name}**!")
    st.caption(f"Roll No: {st.session_state.student_roll.upper()}")
    st.markdown("---")
    st.markdown("#### 📋 Choose a Quiz")

    all_done = all(qn in st.session_state.locked_quizzes for qn in QUIZZES)

    for quiz_name, questions in QUIZZES.items():
        locked = quiz_name in st.session_state.locked_quizzes
        col_a, col_b = st.columns([3, 1])
        with col_a:
            emoji = "🔒" if locked else "📝"
            status = " — **Submitted (Locked)**" if locked else ""
            st.markdown(f"**{emoji} {quiz_name}**{status}")
            if locked:
                st.caption("✅ Already submitted — one attempt allowed per quiz.")
            else:
                st.caption("15 questions • 10 MCQ • 3 Fill in the Blank • 2 Short Answer")
        with col_b:
            if locked:
                score_info = st.session_state.scores.get(quiz_name, {})
                auto  = score_info.get("auto_score", 0)
                total = score_info.get("auto_total", 13)
                st.markdown(
                    f'<div style="background:#1b7f79;color:white;padding:0.5rem 0.8rem;'
                    f'border-radius:8px;text-align:center;font-weight:700;">'
                    f'Score<br>{auto:.1f}/{total}</div>',
                    unsafe_allow_html=True,
                )
            else:
                if st.button("Start →", key=f"start_{quiz_name}"):
                    seed = get_seed(st.session_state.student_roll, quiz_name)
                    st.session_state.active_quiz    = quiz_name
                    st.session_state.shuffled_qs    = shuffle_questions(questions, seed)
                    st.session_state.answers        = {}
                    st.session_state.q_index        = 0
                    st.session_state.submitted      = False
                    st.session_state.confirm_submit = False
                    st.session_state.start_time     = time.time()
                    st.session_state.page           = "quiz"
                    st.rerun()
        st.markdown("---")

    # View results button
    if st.session_state.locked_quizzes:
        if st.button("📊 View My Results", use_container_width=True):
            st.session_state.active_quiz = st.session_state.locked_quizzes[-1]
            st.session_state.page = "result"
            st.rerun()

    # ── INSTRUCTOR PANEL ─────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("🔐 Instructor Panel — Download Results"):
        pwd = st.text_input("Enter instructor password:", type="password", key="inst_pwd")
        if pwd == "mstshoaib2024":   # change this password as needed
            all_results = load_all_results()
            submitted_count = len(all_results)
            total_students  = len(STUDENT_LIST)
            st.success(f"✅ Access granted — {submitted_count}/{total_students} students have submitted.")

            col_x, col_y = st.columns(2)
            with col_x:
                excel_bytes = generate_excel()
                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_bytes,
                    file_name=f"BSAI_Grammar_Quiz_Results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col_y:
                json_str = json.dumps(all_results, indent=2)
                st.download_button(
                    label="📥 Download Raw JSON",
                    data=json_str,
                    file_name="quiz_results_raw.json",
                    mime="application/json",
                    use_container_width=True,
                )

            # Quick table preview
            st.markdown("**Quick Preview:**")
            preview_rows = []
            for roll, sdata in sorted(all_results.items()):
                q1 = sdata["quizzes"].get(list(QUIZZES.keys())[0], {})
                q2 = sdata["quizzes"].get(list(QUIZZES.keys())[1], {}) if len(QUIZZES) > 1 else {}
                preview_rows.append({
                    "Roll": roll.upper(),
                    "Name": sdata.get("name",""),
                    "Quiz1 Auto": f"{q1.get('auto_score','—')}/{q1.get('auto_total','—')}" if q1 else "Not done",
                    "Quiz2 Auto": f"{q2.get('auto_score','—')}/{q2.get('auto_total','—')}" if q2 else "Not done",
                    "Submitted At (Q1)": q1.get("submitted_at","—"),
                })
            import pandas as pd
            if preview_rows:
                st.dataframe(pd.DataFrame(preview_rows), use_container_width=True)
        elif pwd:
            st.error("❌ Wrong password.")


# ═══════════════════════════════════════════════════════════════════════════
#  PAGE: QUIZ
# ═══════════════════════════════════════════════════════════════════════════

elif st.session_state.page == "quiz":
    qs = st.session_state.shuffled_qs
    total_q = len(qs)
    idx = st.session_state.q_index
    quiz_name = st.session_state.active_quiz

    # ── Top bar ─────────────────────────────────────────────────────────
    col_t1, col_t2, col_t3 = st.columns([2, 2, 1])
    with col_t1:
        st.markdown(f"**{quiz_name}**")
    with col_t2:
        elapsed = int(time.time() - (st.session_state.start_time or time.time()))
        m, s = divmod(elapsed, 60)
        st.markdown(f"⏱️ Time: **{m:02d}:{s:02d}**")
    with col_t3:
        answered = len(st.session_state.answers)
        st.markdown(f"✅ **{answered}/{total_q}**")

    progress_bar(idx + 1, total_q)
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Current question ─────────────────────────────────────────────────
    q = qs[idx]
    st.markdown(
        f'<div class="card">'
        f'{type_chip(q["type"])}'
        f'<div class="q-text">Q{idx+1}. {q["text"]}</div>',
        unsafe_allow_html=True,
    )

    key = q["id"]

    if q["type"] == "mcq":
        prev = st.session_state.answers.get(key, None)
        choice = st.radio(
            "Select your answer:",
            q["options"],
            index=q["options"].index(prev) if prev in q["options"] else 0,
            key=f"radio_{key}",
        )
        st.session_state.answers[key] = choice

    elif q["type"] == "fill":
        blanks = q.get("blanks", 1)
        fill_vals = st.session_state.answers.get(key, [""] * blanks)
        if not isinstance(fill_vals, list):
            fill_vals = [fill_vals]
        new_vals = []
        for b in range(blanks):
            val = st.text_input(
                f"Blank {b+1}:",
                value=fill_vals[b] if b < len(fill_vals) else "",
                key=f"fill_{key}_{b}",
            )
            new_vals.append(val)
        st.session_state.answers[key] = new_vals

    elif q["type"] == "short":
        prev_ans = st.session_state.answers.get(key, "")
        ans = st.text_area(
            "Your answer:",
            value=prev_ans,
            height=130,
            key=f"short_{key}",
            placeholder="Write your answer here (2–4 sentences)...",
        )
        st.session_state.answers[key] = ans

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Navigation ──────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    nav1, nav2, nav3 = st.columns([1, 2, 1])

    with nav1:
        if idx > 0:
            if st.button("← Previous"):
                st.session_state.q_index -= 1
                st.rerun()

    with nav2:
        # question number quick-jump
        jump = st.selectbox(
            "Jump to question:",
            options=list(range(1, total_q + 1)),
            index=idx,
            key="jump_select",
        )
        if jump - 1 != idx:
            st.session_state.q_index = jump - 1
            st.rerun()

    with nav3:
        if idx < total_q - 1:
            if st.button("Next →"):
                st.session_state.q_index += 1
                st.rerun()

    st.markdown("---")

    # ── Question overview dots ───────────────────────────────────────────
    st.markdown("**Question Overview:**")
    dots_html = ""
    for i, _q in enumerate(qs):
        answered_flag = _q["id"] in st.session_state.answers and bool(
            st.session_state.answers[_q["id"]]
        )
        color = "#1b7f79" if answered_flag else "#ddd"
        border = "3px solid #e8a020" if i == idx else "2px solid transparent"
        dots_html += (
            f'<span style="display:inline-block;width:28px;height:28px;border-radius:50%;'
            f'background:{color};border:{border};text-align:center;line-height:26px;'
            f'font-size:0.75rem;font-weight:700;color:white;margin:2px;cursor:pointer;">'
            f'{i+1}</span>'
        )
    st.markdown(dots_html, unsafe_allow_html=True)
    st.caption("🟢 Answered &nbsp;&nbsp; ⚪ Not answered &nbsp;&nbsp; 🟡 border = current")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Submit section ───────────────────────────────────────────────────
    unanswered = total_q - len([
        k for k, v in st.session_state.answers.items()
        if v and v != [""] * q.get("blanks", 1)
    ])

    if not st.session_state.confirm_submit:
        col_s1, col_s2 = st.columns([2, 1])
        with col_s1:
            if unanswered > 0:
                st.warning(f"⚠️ You have **{unanswered}** unanswered question(s).")
        with col_s2:
            if st.button("📤 Submit Quiz", use_container_width=True, type="primary"):
                st.session_state.confirm_submit = True
                st.rerun()
    else:
        st.markdown("""
<div class="confirm-box">
  <h3>⚠️ Are you sure you want to submit?</h3>
  <p>Once submitted, you <strong>cannot</strong> change your answers.<br>
  Make sure you have reviewed all your responses.</p>
</div>
""", unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        cc1, cc2, cc3 = st.columns([1, 1, 1])
        with cc1:
            if st.button("✅ Yes, Submit Now", use_container_width=True, type="primary"):
                # ── Calculate score ──────────────────────────────────────
                auto_score = 0.0
                auto_total = 0
                detail = []
                for _q in qs:
                    _key = _q["id"]
                    _ans = st.session_state.answers.get(_key, "")
                    if _q["type"] == "mcq":
                        s = score_mcq(_q, _ans if _ans else "")
                        auto_score += s
                        auto_total += 1
                        detail.append({"q": _q, "ans": _ans, "score": s, "max": 1})
                    elif _q["type"] == "fill":
                        _fill_ans = _ans if isinstance(_ans, list) else [_ans]
                        s = score_fill(_q, _fill_ans)
                        auto_score += s
                        auto_total += 1
                        detail.append({"q": _q, "ans": _fill_ans, "score": s, "max": 1})
                    elif _q["type"] == "short":
                        detail.append({"q": _q, "ans": _ans, "score": None, "max": 2})

                elapsed_final = int(time.time() - (st.session_state.start_time or time.time()))
                score_payload = {
                    "auto_score":   auto_score,
                    "auto_total":   auto_total,
                    "detail":       detail,
                    "elapsed":      elapsed_final,
                    "submitted_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                }
                st.session_state.scores[quiz_name] = score_payload
                if quiz_name not in st.session_state.completed_quizzes:
                    st.session_state.completed_quizzes.append(quiz_name)
                if quiz_name not in st.session_state.locked_quizzes:
                    st.session_state.locked_quizzes.append(quiz_name)
                # ── Persist to disk — one-time lock ──────────────────────
                save_result(
                    st.session_state.student_roll,
                    st.session_state.student_name,
                    quiz_name,
                    score_payload,
                )
                st.session_state.submitted = True
                st.session_state.page = "result"
                st.rerun()

        with cc3:
            if st.button("❌ Cancel, Go Back", use_container_width=True):
                st.session_state.confirm_submit = False
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════
#  PAGE: RESULT
# ═══════════════════════════════════════════════════════════════════════════

elif st.session_state.page == "result":
    quiz_name = st.session_state.active_quiz or (
        st.session_state.completed_quizzes[-1] if st.session_state.completed_quizzes else None
    )

    if quiz_name and quiz_name in st.session_state.scores:
        score_data = st.session_state.scores[quiz_name]
        auto = score_data["auto_score"]
        auto_total = score_data["auto_total"]
        elapsed = score_data.get("elapsed", 0)
        m, s = divmod(elapsed, 60)
        pct = int((auto / auto_total) * 100) if auto_total > 0 else 0

        grade = "A+" if pct >= 90 else "A" if pct >= 80 else "B" if pct >= 70 else "C" if pct >= 60 else "D" if pct >= 50 else "F"
        grade_color = "#27ae60" if pct >= 70 else "#f39c12" if pct >= 50 else "#e74c3c"

        st.markdown(f"""
<div class="score-box">
  <div style="font-size:1rem;opacity:0.8;margin-bottom:0.5rem;">{st.session_state.student_name} &nbsp;•&nbsp; {st.session_state.student_roll.upper()}</div>
  <div style="font-size:1.2rem;margin-bottom:0.5rem;">{quiz_name}</div>
  <div class="big-score">{auto:.0f}<span style="font-size:2rem;">/{auto_total}</span></div>
  <div class="score-label">Auto-graded score (MCQ + Fill in the Blank)</div>
  <div style="margin-top:1rem;">
    <span style="background:{grade_color};padding:0.4rem 1.5rem;border-radius:20px;font-weight:700;font-size:1.5rem;">{grade}</span>
    &nbsp;&nbsp;
    <span style="opacity:0.8;">{pct}% &nbsp;|&nbsp; ⏱️ {m:02d}:{s:02d}</span>
  </div>
  <div style="font-size:0.82rem;opacity:0.6;margin-top:0.7rem;">Submitted: {score_data.get("submitted_at","")}<br>Short answer questions (2 marks each) to be graded by instructor.</div>
</div>
""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Detailed review ──────────────────────────────────────────────
        with st.expander("📖 Detailed Answer Review", expanded=True):
            for i, item in enumerate(score_data["detail"]):
                q = item["q"]
                ans = item["ans"]
                sc = item["score"]

                if q["type"] == "mcq":
                    correct = sc == 1.0
                    css_cls = "ans-correct" if correct else "ans-wrong"
                    mark = "✅" if correct else "❌"
                    st.markdown(
                        f'<div class="{css_cls}"><b>{mark} Q{i+1} [MCQ]:</b> {q["text"]}<br>'
                        f'<b>Your answer:</b> {ans}<br>'
                        f'<b>Correct answer:</b> {q["answer"]}</div>',
                        unsafe_allow_html=True,
                    )

                elif q["type"] == "fill":
                    correct_list = q["answer"]
                    your_list = ans if isinstance(ans, list) else [ans]
                    all_correct = all(
                        (your_list[j].strip().lower() == correct_list[j].strip().lower() if j < len(your_list) else False)
                        for j in range(len(correct_list))
                    )
                    css_cls = "ans-correct" if all_correct else ("ans-partial" if sc > 0 else "ans-wrong")
                    mark = "✅" if all_correct else ("⚠️" if sc > 0 else "❌")
                    your_str = " | ".join(your_list) if your_list else "(blank)"
                    correct_str = " | ".join(correct_list)
                    st.markdown(
                        f'<div class="{css_cls}"><b>{mark} Q{i+1} [Fill]:</b> {q["text"]}<br>'
                        f'<b>Your answer(s):</b> {your_str}<br>'
                        f'<b>Expected:</b> {correct_str}</div>',
                        unsafe_allow_html=True,
                    )

                elif q["type"] == "short":
                    st.markdown(
                        f'<div class="ans-partial"><b>📝 Q{i+1} [Short — Pending Manual Grading]:</b> {q["text"]}<br>'
                        f'<b>Your answer:</b> {ans if ans else "(no answer)"}<br>'
                        f'<b>Model answer:</b> {q["answer"]}</div>',
                        unsafe_allow_html=True,
                    )

    # ── Navigation ───────────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    bc1, bc2 = st.columns(2)
    with bc1:
        if st.button("← Back to Quiz Portal", use_container_width=True):
            st.session_state.page = "quiz_select"
            st.session_state.confirm_submit = False
            st.rerun()
    with bc2:
        other_quizzes = [q for q in QUIZZES if q not in st.session_state.completed_quizzes]
        if other_quizzes:
            if st.button(f"📝 Take Next Quiz →", use_container_width=True):
                next_q = other_quizzes[0]
                seed = get_seed(st.session_state.student_roll, next_q)
                st.session_state.active_quiz = next_q
                st.session_state.shuffled_qs = shuffle_questions(QUIZZES[next_q], seed)
                st.session_state.answers = {}
                st.session_state.q_index = 0
                st.session_state.submitted = False
                st.session_state.confirm_submit = False
                st.session_state.start_time = time.time()
                st.session_state.page = "quiz"
                st.rerun()


# ── Footer ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
  Modern Grammar Quiz Portal &nbsp;•&nbsp; BS AI SP26 &nbsp;•&nbsp; COMSATS University Lahore<br>
  Designed by <b>Muhammad Shoaib Tahir</b> &nbsp;•&nbsp; Based on <i>Modern Grammar Guide</i>
</div>
""", unsafe_allow_html=True)